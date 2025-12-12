import win32com.client
import sys
import time
import pandas as pd
import os
import pyperclip
import pythoncom

class SapBotConversiones:
    def __init__(self):
        self.session = None
        # Inicializar COM para este hilo
        try:
            pythoncom.CoInitialize()
        except:
            pass
        self.connect_to_sap()

    def connect_to_sap(self):
        try:
            sap_gui = win32com.client.GetObject("SAPGUI")
            if not sap_gui:
                print("No se encontró SAPGUI. Intentando Dispatch...")
                sap_gui = win32com.client.Dispatch("SAPGUI")
            
            application = sap_gui.GetScriptingEngine
            connection = application.Children(0)
            self.session = connection.Children(0)
            print("--- Conectado a SAP exitosamente ---")
        except Exception as e:
            print(f"Error conectando a SAP: {e}")
            print("Asegúrate de tener SAP abierto y logueado.")

    def obtener_conversiones_mm03(self, materiales):
        """
        Obtiene factores de conversión usando MM03 (material por material)
        materiales: lista de códigos de material
        """
        print("--- Consultando Conversiones con MM03 ---")
        
        conversiones = []
        
        for idx, material in enumerate(materiales):
            print(f"[{idx+1}/{len(materiales)}] Procesando material: {material}")
            
            try:
                # Ir a MM03
                print(f"   -> Abriendo MM03...")
                self.session.findById("wnd[0]/tbar[0]/okcd").text = "/nMM03"
                self.session.findById("wnd[0]").sendVKey(0)
                time.sleep(1.0)
                
                # Ingresar material
                print(f"   -> Ingresando material {material}...")
                self.session.findById("wnd[0]/usr/ctxtRMMG1-MATNR").text = material
                self.session.findById("wnd[0]").sendVKey(0)
                time.sleep(1.0)
                
                # Seleccionar vistas (Datos básicos 1 y 2)
                try:
                    print(f"   -> Seleccionando vistas...")
                    # Seleccionar primera y segunda vista (Datos básicos)
                    self.session.findById("wnd[1]/usr/tblSAPLMGMMTC_VIEW").getAbsoluteRow(0).selected = True
                    self.session.findById("wnd[1]/usr/tblSAPLMGMMTC_VIEW").getAbsoluteRow(1).selected = True
                    self.session.findById("wnd[1]/tbar[0]/btn[0]").press()
                    time.sleep(1.0)
                except Exception as e:
                    # Si no hay popup de selección, continuar
                    print(f"   -> Sin popup de vistas (normal): {e}")
                    pass
                
                # Ir a pestaña de Unidades de medida alternativas
                try:
                    print(f"   -> Buscando pestaña de unidades alternativas...")
                    self.session.findById("wnd[0]/usr/tabsTABSPR1/tabpZU02").select()
                    time.sleep(1.0)
                    print(f"   -> Pestaña UMV encontrada")
                except Exception as e:
                    print(f"   [WARN] No se pudo acceder a pestaña UMV: {e}")
                    print(f"   [DEBUG] Intentando salir de MM03...")
                    try:
                        self.session.findById("wnd[0]/tbar[0]/btn[12]").press()  # Salir
                    except:
                        pass
                    continue
                
                # Leer tabla de unidades alternativas
                try:
                    print(f"   -> Leyendo tabla de conversiones...")
                    tabla = self.session.findById("wnd[0]/usr/tabsTABSPR1/tabpZU02/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:8020/tblSAPLMGD1TC_ME_8020")
                    row_count = tabla.RowCount
                    
                    print(f"   -> Encontradas {row_count} unidades alternativas")
                    
                    for r in range(row_count):
                        try:
                            # Leer valores de la fila
                            # Columnas típicas: UMV, Numerador, Denominador
                            umv = tabla.GetCell(r, 5).Text.strip()  # Unidad alternativa
                            
                            # Intentar leer numerador y denominador
                            try:
                                numerador = tabla.GetCell(r, 1).Text.strip()
                                denominador = tabla.GetCell(r, 2).Text.strip()
                            except:
                                numerador = "1"
                                denominador = "1"
                            
                            if umv and umv not in ["", "UN"]:  # Ignorar unidad base
                                conversiones.append({
                                    "Material": material,
                                    "UMV": umv,
                                    "Numerador": numerador,
                                    "Denominador": denominador
                                })
                                print(f"      {umv}: {numerador}/{denominador}")
                        except Exception as e:
                            # Fila vacía o error de lectura
                            continue
                            
                except Exception as e:
                    print(f"   [ERROR] No se pudo leer tabla de UMV: {e}")
                    continue
                
                # Salir de MM03 para el siguiente material
                try:
                    self.session.findById("wnd[0]/tbar[0]/btn[12]").press()  # Salir
                    time.sleep(0.5)
                except:
                    pass
                
            except Exception as e:
                print(f"   [ERROR] Error procesando material {material}: {e}")
                # Intentar salir de cualquier pantalla
                try:
                    self.session.findById("wnd[0]/tbar[0]/btn[12]").press()
                except:
                    pass
                continue
        
        df_conversiones = pd.DataFrame(conversiones)
        print(f"\n✅ Total conversiones obtenidas: {len(df_conversiones)}")
        return df_conversiones

    def run(self, excel_path):
        """
        excel_path: path al archivo Excel con materiales
        """
        if not os.path.exists(excel_path):
            print(f"Archivo no encontrado: {excel_path}")
            return
        
        print(f"Leyendo materiales desde: {excel_path}")
        
        # Leer materiales del Excel (buscar hoja Hoja1 o la primera hoja que no sea Lx02/Conversiones)
        df_input = None
        try:
            # Intentar leer "Hoja1" específicamente
            df_input = pd.read_excel(excel_path, sheet_name="Hoja1", dtype=str)
            print("   Leyendo desde hoja: Hoja1")
        except:
            # Si no existe Hoja1, buscar la primera hoja que no sea Lx02 o Conversiones
            try:
                xl_file = pd.ExcelFile(excel_path)
                for sheet_name in xl_file.sheet_names:
                    if sheet_name.lower() not in ["lx02", "conversiones"]:
                        df_input = pd.read_excel(excel_path, sheet_name=sheet_name, dtype=str)
                        print(f"   Leyendo desde hoja: {sheet_name}")
                        break
            except:
                # Último recurso: primera hoja
                df_input = pd.read_excel(excel_path, sheet_name=0, dtype=str)
                print("   Leyendo desde primera hoja")
        
        if df_input is None or df_input.empty:
            print("❌ No se pudo leer el archivo Excel.")
            return
        
        # Buscar columna Material (flexible: Material, SKU, Código, etc.)
        mat_col = None
        posibles_nombres = ["material", "sku", "codigo", "código", "mat", "matnr"]
        
        for col in df_input.columns:
            col_lower = col.lower().strip()
            for nombre in posibles_nombres:
                if nombre in col_lower:
                    mat_col = col
                    print(f"   Columna de materiales detectada: '{col}'")
                    break
            if mat_col:
                break
        
        if not mat_col:
            print(f"❌ No se encontró columna de materiales.")
            print(f"   Columnas disponibles: {df_input.columns.tolist()}")
            print(f"   Buscando: {posibles_nombres}")
            return
        
        materiales_requeridos = df_input[mat_col].dropna().unique().tolist()
        print(f"Materiales en archivo: {len(materiales_requeridos)}")
        
        # Intentar leer conversiones existentes de la pestaña "Conversiones"
        df_conversiones_existentes = pd.DataFrame()
        try:
            df_conversiones_existentes = pd.read_excel(excel_path, sheet_name="Conversiones", dtype=str)
            print(f"✅ Conversiones existentes encontradas: {len(df_conversiones_existentes)} registros")
        except:
            print("ℹ️  No se encontró pestaña 'Conversiones'. Se creará una nueva.")
        
        # Determinar qué materiales faltan
        if not df_conversiones_existentes.empty and "Material" in df_conversiones_existentes.columns:
            materiales_con_conversion = df_conversiones_existentes["Material"].unique().tolist()
            materiales_faltantes = [m for m in materiales_requeridos if m not in materiales_con_conversion]
        else:
            materiales_faltantes = materiales_requeridos
        
        print(f"Materiales sin conversión: {len(materiales_faltantes)}")
        
        # Consultar SAP solo para materiales faltantes
        df_nuevas_conversiones = pd.DataFrame()
        if materiales_faltantes:
            print(f"\n🔍 Consultando SAP para {len(materiales_faltantes)} materiales...")
            df_nuevas_conversiones = self.obtener_conversiones_mm03(materiales_faltantes)
        else:
            print("\n✅ Todos los materiales ya tienen conversiones en el Excel.")
        
        # Combinar conversiones existentes con nuevas
        if not df_conversiones_existentes.empty and not df_nuevas_conversiones.empty:
            df_conversiones_final = pd.concat([df_conversiones_existentes, df_nuevas_conversiones], ignore_index=True)
        elif not df_nuevas_conversiones.empty:
            df_conversiones_final = df_nuevas_conversiones
        else:
            df_conversiones_final = df_conversiones_existentes
        
        if df_conversiones_final.empty:
            print("No se obtuvieron conversiones.")
            return
        
        # Guardar en la pestaña "Conversiones" del mismo Excel
        print(f"\n💾 Guardando conversiones en pestaña 'Conversiones'...")
        try:
            # Abrir Excel con openpyxl para preservar otras hojas
            from openpyxl import load_workbook
            
            # Cargar workbook existente
            wb = load_workbook(excel_path)
            
            # Eliminar hoja "Conversiones" si existe
            if "Conversiones" in wb.sheetnames:
                del wb["Conversiones"]
            
            # Crear nueva hoja "Conversiones"
            wb.create_sheet("Conversiones")
            
            # Guardar workbook
            wb.save(excel_path)
            
            # Ahora escribir el DataFrame en la hoja
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df_conversiones_final.to_excel(writer, sheet_name='Conversiones', index=False)
            
            print(f"✅ Conversiones guardadas: {len(df_conversiones_final)} registros totales")
            print(f"   - Existentes: {len(df_conversiones_existentes)}")
            print(f"   - Nuevas: {len(df_nuevas_conversiones)}")
            
        except Exception as e:
            print(f"❌ Error guardando en Excel: {e}")
            # Fallback: guardar en archivo separado
            output_path = excel_path.replace(".xlsx", "_conversiones.xlsx")
            df_conversiones_final.to_excel(output_path, index=False)
            print(f"   Conversiones guardadas en: {output_path}")
        
        return df_conversiones_final


if __name__ == "__main__":
    # Uso: El bot lee materiales del Excel, consulta solo los que faltan,
    # y guarda las conversiones en una pestaña "Conversiones" del mismo archivo
    
    bot = SapBotConversiones()
    
    excel_path = r"C:\Users\ariel.mella\OneDrive - CIAL Alimentos\Archivos de Operación  Outbound CD - 16.-Inventario Critico\carga_Lt01.xlsx"
    bot.run(excel_path)
